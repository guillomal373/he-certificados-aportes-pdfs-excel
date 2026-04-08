#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SHEET_COLUMNS = {
    "Liquidaciones Pagadas": [
        "archivo_pdf",
        "nombre_persona",
        "tipo_id",
        "numero_id",
        "periodo_pension",
        "periodo_salud",
        "tipo_planilla",
        "clave",
        "no_transaccion",
        "fecha_pago",
        "fecha_generacion_certificado",
    ],
    "Seguridad Social": [
        "archivo_pdf",
        "nombre_persona",
        "tipo_id",
        "numero_id",
        "periodo_pension",
        "tipo_planilla",
        "afp_administradora",
        "afp_dias",
        "afp_ibc",
        "afp_fs",
        "afp_fsp",
        "afp_aporte",
        "eps_administradora",
        "eps_dias",
        "eps_ibc",
        "eps_aporte",
        "arl_administradora",
        "arl_dias",
        "arl_ibc",
        "arl_aporte",
        "fecha_generacion_certificado",
    ],
    "Aportes Parafiscales": [
        "archivo_pdf",
        "nombre_persona",
        "tipo_id",
        "numero_id",
        "periodo_pension",
        "tipo_planilla",
        "ccf_administradora",
        "ccf_dias",
        "ccf_ibc",
        "ccf_aporte",
        "otros_parafiscales_ibc",
        "icbf_tarifa",
        "icbf_aporte",
        "sena_tarifa",
        "sena_aporte",
        "esap_tarifa",
        "esap_aporte",
        "men_tarifa",
        "men_aporte",
        "fecha_generacion_certificado",
    ],
    "Novedades": [
        "archivo_pdf",
        "nombre_persona",
        "tipo_id",
        "numero_id",
        "periodo_pension",
        "tipo_planilla",
        "codigo_novedad",
        "marcado",
        "fecha_generacion_certificado",
    ],
}

NOVEDADES_CODIGOS = [
    "ING", "RET", "TAE", "TDE", "TAP",
    "TDP", "VSP", "VST", "SLN", "IGE",
    "LMA", "VAC", "AVP", "IRL", "VCT",
]

SECURITY_PATTERN = re.compile(
    r"(\d{4}/\d{2})\s+([A-Z])\s+([A-Z0-9ÁÉÍÓÚÜÑ()./\- ]+?)\s+(\d+)\s+\$([\d,]+)\s+\$([\d,]+)\s+\$([\d,]+)\s+\$([\d,]+)\s+"
    r"([A-Z0-9ÁÉÍÓÚÜÑ()./\- ]+?)\s+(\d+)\s+\$([\d,]+)\s+\$([\d,]+)\s+([A-Z0-9ÁÉÍÓÚÜÑ()./\- ]+?)\s+(\d+)\s+\$([\d,]+)\s+\$([\d,]+)"
)

PARAF_PATTERN = re.compile(
    r"(\d{4}/\d{2})\s+([A-Z])\s+([A-Z0-9ÁÉÍÓÚÜÑ()./\- ]+?)\s+(\d+)\s+\$([\d,]+)\s+\$([\d,]+)\s+\$([\d,]+)\s+"
    r"(\d+)%\s+\$([\d,]+)\s+(\d+)%\s+\$([\d,]+)\s+(\d+)%\s+\$([\d,]+)\s+(\d+)%\s+\$([\d,]+)"
)

LIQ_PATTERN = re.compile(
    r"(\d{4}/\d{2})\s+(\d{4}/\d{2})\s+([A-Z])\s+(\d+)\s+(\d+)\s+(\d{4}/\d{2}/\d{2})"
)

HEADER_PATTERN = re.compile(
    r"Se certifica que\s+.*?para\s+(.+?)\s+identificado con\s+([A-Z]{2})\s+(\d+)\s*:?",
    re.IGNORECASE | re.DOTALL,
)

GEN_DATE_PATTERN = re.compile(
    r"Certificado generado el\s+(\d{4}-\d{2}-\d{2})\s+a las\s+(\d{2}:\d{2})",
    re.IGNORECASE,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Procesa certificados de aportes PDF en una carpeta y genera un Excel consolidado."
    )
    parser.add_argument("--input", required=True, help="Carpeta que contiene los PDF.")
    parser.add_argument("--output", required=True, help="Ruta completa del XLSX de salida.")
    return parser.parse_args()


def clean_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()

def normalize_security_section_text(text: str) -> str:
    replacements = {
        "EPS SURA (ANTES\nSUSALUD)": "EPS SURA (ANTES SUSALUD)",
        "EPS SURA\n(ANTES SUSALUD)": "EPS SURA (ANTES SUSALUD)",
        "EPS SURA(ANTES\nSUSALUD)": "EPS SURA (ANTES SUSALUD)",
        "EPS SURA(ANTES SUSALUD)": "EPS SURA (ANTES SUSALUD)",
        "EPSSURA(ANTES\nSUSALUD)": "EPS SURA (ANTES SUSALUD)",
        "EPSSURA(ANTES SUSALUD)": "EPS SURA (ANTES SUSALUD)",
        "EPSSURA(ANTES": "EPS SURA (ANTES",
        "EPSSURA(ANTES": "EPS SURA (ANTES",
        "COLPATRIA\nARP": "COLPATRIA ARP",
        "COLPATRIAARP": "COLPATRIA ARP",
        "ARL\nSURA": "ARL SURA",
        "ARLSURA": "ARL SURA",
        "SALUD\nTOTAL": "SALUD TOTAL",
        "SALUDTOTAL": "SALUD TOTAL",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    return clean_spaces(text)

def normalize_parafiscales_section_text(text: str) -> str:
    replacements = {
        "COLSUBSIDIO": "COLSUBSIDIO",
        "COLSUBSIDIO\n": "COLSUBSIDIO ",
        "CAFAM\n": "CAFAM ",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    return clean_spaces(text)

def money_to_int(value: str) -> int:
    return int(re.sub(r"[^\d]", "", value))


def parse_date(value: str) -> datetime:
    return datetime.strptime(value.replace("-", "/"), "%Y/%m/%d")


def normalize_admin_name(value: str) -> str:
    raw = clean_spaces(value.upper())

    raw = raw.replace("ARLSURA", "ARL SURA")
    raw = raw.replace("COLPATRIAARP", "COLPATRIA ARP")
    raw = raw.replace("SALUDTOTAL", "SALUD TOTAL")
    raw = raw.replace("EPSSURA", "EPS SURA")
    raw = raw.replace("EPSSURA", "EPS SURA")

    raw = raw.replace("EPS SURA(ANTES SUSALUD)", "EPS SURA (ANTES SUSALUD)")
    raw = raw.replace("EPS SURA(ANTES", "EPS SURA (ANTES")
    raw = raw.replace("EPS SURA (ANTES", "EPS SURA (ANTES SUSALUD)")

    raw = clean_spaces(raw)

    return raw


def crop_text(page) -> str:
    candidates = []

    try:
        text_full = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
        candidates.append(text_full)
    except Exception:
        pass

    try:
        cropped1 = page.crop((0, 100, page.width, page.height - 20))
        text_c1 = cropped1.extract_text(x_tolerance=2, y_tolerance=3) or ""
        candidates.append(text_c1)
    except Exception:
        pass

    try:
        cropped2 = page.crop((0, 60, page.width, page.height - 20))
        text_c2 = cropped2.extract_text(x_tolerance=2, y_tolerance=3) or ""
        candidates.append(text_c2)
    except Exception:
        pass

    try:
        words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
        text_words = " ".join(w["text"] for w in words)
        candidates.append(text_words)
    except Exception:
        pass

    normalized = [clean_spaces(t) for t in candidates if t and t.strip()]

    for text in normalized:
        if "Se certifica que" in text:
            return text[text.find("Se certifica que"):]

    if normalized:
        return max(normalized, key=len)

    return ""


def extract_full_text(pdf) -> str:
    page_texts = []

    for i, page in enumerate(pdf.pages):
        text = crop_text(page)
        if text:
            page_texts.append(text)

    full_text = clean_spaces(" ".join(page_texts))
    return full_text


def extract_tables_from_all_pages(pdf):
    all_tables = []

    for page in pdf.pages:
        try:
            tables = page.extract_tables()
            if tables:
                all_tables.extend(tables)
        except Exception:
            pass

    return all_tables


def extract_last_identity(full_text: str) -> Tuple[str, str, int]:
    matches = HEADER_PATTERN.findall(full_text)
    if not matches:
        raise ValueError(f"No se pudo extraer la identidad del encabezado. Texto detectado: {full_text[:500]}")
    person_name, id_type, id_number = matches[-1]
    return clean_spaces(person_name), id_type.upper(), int(id_number)


def extract_generation_date(full_text: str) -> datetime:
    match = GEN_DATE_PATTERN.search(full_text)
    if not match:
        raise ValueError("No se pudo extraer la fecha de generación.")
    date_part, time_part = match.groups()
    return datetime.strptime(f"{date_part} {time_part}", "%Y-%m-%d %H:%M")


def get_section(full_text: str, start: str, end: str | None) -> str:
    start_idx = full_text.find(start)
    if start_idx == -1:
        return ""
    if end is None:
        return full_text[start_idx:]
    end_idx = full_text.find(end, start_idx)
    if end_idx == -1:
        return full_text[start_idx:]
    return full_text[start_idx:end_idx]


def extract_tables(pdf):
    tables = extract_tables_from_all_pages(pdf)
    return tables

def parse_liquidaciones(
    section_text: str,
    archivo_pdf: str,
    nombre: str,
    tipo_id: str,
    numero_id: int,
    fecha_generacion: datetime,
) -> List[List]:
    rows = []
    for periodo_pension, periodo_salud, tipo_planilla, clave, no_transaccion, fecha_pago in LIQ_PATTERN.findall(section_text):
        rows.append([
            archivo_pdf,
            nombre,
            tipo_id,
            numero_id,
            periodo_pension,
            periodo_salud,
            tipo_planilla,
            int(clave),
            int(no_transaccion),
            parse_date(fecha_pago),
            fecha_generacion,
        ])
    return rows

def parse_seguridad_social(
    section_text: str,
    archivo_pdf: str,
    nombre: str,
    tipo_id: str,
    numero_id: int,
    fecha_generacion: datetime,
) -> List[List]:
    rows = []

    print(f"DEBUG SEGURIDAD SOCIAL - ARCHIVO: {archivo_pdf}")

    section_text = normalize_security_section_text(section_text)

    # print("TEXTO NORMALIZADO SEGURIDAD SOCIAL:")
    # print(section_text[:2000])
    # print("-" * 80)

    matches = SECURITY_PATTERN.findall(section_text)

    print(f"TOTAL MATCHES SEGURIDAD SOCIAL: {len(matches)}")

    if not matches:
        # print(f"ERROR DEBUG - No hubo matches en Seguridad Social para {archivo_pdf}")
        print("=" * 80)
        return rows

    for i, match in enumerate(matches, start=1):
        # print(f"MATCH #{i}: {match}")

        (
            periodo_pension, tipo_planilla,
            afp_adm, afp_dias, afp_ibc, afp_fs, afp_fsp, afp_aporte,
            eps_adm, eps_dias, eps_ibc, eps_aporte,
            arl_adm, arl_dias, arl_ibc, arl_aporte,
        ) = match

        afp_adm_norm = normalize_admin_name(afp_adm)
        eps_adm_norm = normalize_admin_name(eps_adm)
        arl_adm_norm = normalize_admin_name(arl_adm)

        # print(f"DEBUG NORMALIZADO -> AFP:[{afp_adm_norm}] EPS:[{eps_adm_norm}] ARL:[{arl_adm_norm}]")

        rows.append([
            archivo_pdf,
            nombre,
            tipo_id,
            numero_id,
            periodo_pension,
            tipo_planilla,
            afp_adm_norm,
            int(afp_dias),
            money_to_int(afp_ibc),
            money_to_int(afp_fs),
            money_to_int(afp_fsp),
            money_to_int(afp_aporte),
            eps_adm_norm,
            int(eps_dias),
            money_to_int(eps_ibc),
            money_to_int(eps_aporte),
            arl_adm_norm,
            int(arl_dias),
            money_to_int(arl_ibc),
            money_to_int(arl_aporte),
            fecha_generacion,
        ])

    print(f"TOTAL FILAS GENERADAS SEGURIDAD SOCIAL: {len(rows)}")
    print("=" * 80)

    return rows

def parse_parafiscales(
    section_text: str,
    archivo_pdf: str,
    nombre: str,
    tipo_id: str,
    numero_id: int,
    fecha_generacion: datetime,
) -> List[List]:
    rows = []
    section_text = normalize_parafiscales_section_text(section_text)
    
    for match in PARAF_PATTERN.findall(section_text):
        (
            periodo_pension, tipo_planilla,
            ccf_adm, ccf_dias, ccf_ibc, ccf_aporte, otros_ibc,
            icbf_tarifa, icbf_aporte,
            sena_tarifa, sena_aporte,
            esap_tarifa, esap_aporte,
            men_tarifa, men_aporte,
        ) = match

        rows.append([
            archivo_pdf,
            nombre,
            tipo_id,
            numero_id,
            periodo_pension,
            tipo_planilla,
            normalize_admin_name(ccf_adm),
            int(ccf_dias),
            money_to_int(ccf_ibc),
            money_to_int(ccf_aporte),
            money_to_int(otros_ibc),
            int(icbf_tarifa),
            money_to_int(icbf_aporte),
            int(sena_tarifa),
            money_to_int(sena_aporte),
            int(esap_tarifa),
            money_to_int(esap_aporte),
            int(men_tarifa),
            money_to_int(men_aporte),
            fecha_generacion,
        ])
    return rows


def parse_novedades_from_table(
    table,
    archivo_pdf: str,
    nombre: str,
    tipo_id: str,
    numero_id: int,
    fecha_generacion: datetime,
) -> List[List]:
    rows = []

    if not table or len(table) < 3:
        return rows

    data_rows = table[2:]

    for row in data_rows:
        if not row:
            continue

        row = [(cell or "").strip() for cell in row]

        if len(row) < 17:
            continue

        periodo_match = re.search(r"\d{4}/\d{2}", row[0] or "")
        tipo_match = re.search(r"\b[A-Z]\b", clean_spaces(row[1] or ""))

        if not periodo_match or not tipo_match:
            continue

        periodo_pension = periodo_match.group(0)
        tipo_planilla = tipo_match.group(0)

        for idx, codigo in enumerate(NOVEDADES_CODIGOS, start=2):
            if idx >= len(row):
                continue

            cell_value = clean_spaces((row[idx] or "").upper())
            if "X" in cell_value:
                rows.append([
                    archivo_pdf,
                    nombre,
                    tipo_id,
                    numero_id,
                    periodo_pension,
                    tipo_planilla,
                    codigo,
                    "X",
                    fecha_generacion,
                ])

    return rows


def process_pdf(pdf_path: Path) -> Dict[str, List[List]]:
    with pdfplumber.open(pdf_path) as pdf:
        full_text = extract_full_text(pdf)
        print(f"DEBUG TEXTO {pdf_path.name}:")
        print(full_text[:800])
        print("-" * 80)

        tables = extract_tables(pdf)

    nombre, tipo_id, numero_id = extract_last_identity(full_text)
    fecha_generacion = extract_generation_date(full_text)

    section_liq = get_section(
        full_text,
        "Datos de las Liquidaciones Pagadas",
        "Aportes Sistema de Seguridad Social",
    )
    section_seg = get_section(
        full_text,
        "Aportes Sistema de Seguridad Social",
        "Aportes Parafiscales",
    )
    section_par = get_section(
        full_text,
        "Aportes Parafiscales",
        "Novedades",
    )

    novedades_table = tables[-1] if tables else []

    data = {
        "Liquidaciones Pagadas": parse_liquidaciones(
            section_liq, pdf_path.name, nombre, tipo_id, numero_id, fecha_generacion
        ),
        "Seguridad Social": parse_seguridad_social(
            section_seg, pdf_path.name, nombre, tipo_id, numero_id, fecha_generacion
        ),
        "Aportes Parafiscales": parse_parafiscales(
            section_par, pdf_path.name, nombre, tipo_id, numero_id, fecha_generacion
        ),
        "Novedades": parse_novedades_from_table(
            novedades_table, pdf_path.name, nombre, tipo_id, numero_id, fecha_generacion
        ),
    }

    return data


def build_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for sheet_name, columns in SHEET_COLUMNS.items():
        ws = wb.create_sheet(sheet_name)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(col_name) + 3, 14), 28)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
        ws.row_dimensions[1].height = 24

    return wb


def apply_formats(ws, sheet_name: str) -> None:
    if ws.max_row <= 1:
        return

    date_columns_by_sheet = {
        "Liquidaciones Pagadas": {10, 11},
        "Seguridad Social": {21},
        "Aportes Parafiscales": {20},
        "Novedades": {9},
    }

    numeric_columns_by_sheet = {
        "Liquidaciones Pagadas": {4, 8, 9},
        "Seguridad Social": {4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20},
        "Aportes Parafiscales": {4, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19},
        "Novedades": {4},
    }

    date_cols = date_columns_by_sheet[sheet_name]
    numeric_cols = numeric_columns_by_sheet[sheet_name]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column in date_cols and cell.value is not None:
                if isinstance(cell.value, datetime) and cell.column in {11, 20, 21, 9}:
                    cell.number_format = "yyyy-mm-dd hh:mm"
                else:
                    cell.number_format = "yyyy-mm-dd"
            elif cell.column in numeric_cols and cell.value is not None:
                cell.number_format = "0"


def main() -> None:
    args = parse_args()
    input_dir = Path(args.input)
    output_file = Path(args.output)

    if not input_dir.exists() or not input_dir.is_dir():
        raise SystemExit(f"La carpeta de entrada no existe o no es válida: {input_dir}")

    pdf_files = sorted(input_dir.glob("*.pdf"))
    if not pdf_files:
        raise SystemExit(f"No se encontraron archivos PDF en: {input_dir}")

    consolidated = {sheet_name: [] for sheet_name in SHEET_COLUMNS}

    ok_files = []
    error_files = []

    print(f"Procesando {len(pdf_files)} archivos PDF...")

    for pdf_path in pdf_files:
        try:
            data = process_pdf(pdf_path)
            for sheet_name, rows in data.items():
                consolidated[sheet_name].extend(rows)
            ok_files.append(pdf_path.name)
            print(f"OK  - {pdf_path.name}")
        except Exception as exc:
            error_files.append((pdf_path.name, str(exc)))
            print(f"ERROR - {pdf_path.name}: {exc}")

    wb = build_workbook()

    for sheet_name, rows in consolidated.items():
        ws = wb[sheet_name]
        for row in rows:
            ws.append(row)
        apply_formats(ws, sheet_name)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)

    print("\n" + "=" * 80)
    print("RESUMEN FINAL")
    print("=" * 80)
    print(f"Total de PDFs encontrados : {len(pdf_files)}")
    print(f"Procesados OK             : {len(ok_files)}")
    print(f"Con error                 : {len(error_files)}")
    print(f"Excel generado en         : {output_file}")

    if ok_files:
        print("\nArchivos OK:")
        for name in ok_files:
            print(f"  - {name}")

    if error_files:
        print("\nArchivos con ERROR:")
        for name, err in error_files:
            print(f"  - {name}: {err}")


if __name__ == "__main__":
    main()